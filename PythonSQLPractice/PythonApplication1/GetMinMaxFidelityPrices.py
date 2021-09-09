from datetime import date
from openpyxl import load_workbook
from pandas.tseries.offsets import BMonthEnd
import time
from bs4 import BeautifulSoup
import requests
import json
import os

def GetData(URL_Price):
     r=''
     while r == '':
        try:
            r = requests.get(URL_Price)
            encodedText = r.text.encode("utf-8")
            soup = BeautifulSoup(encodedText)
            rows=soup.currentTag.currentTag.currentTag.currentTag.text
            data = json.loads(rows)['TimeSeries']['Security'][0]['HistoryDetail']
            break
        except:
            print("Connection refused by the server..")
            print("Let me sleep for 5 seconds")
            print("ZZzzzz...")
            time.sleep(5)
            print("Was a nice sleep, now let me continue...")
            continue
     return data

def FindSmallestPrice(data):
    price=[]  
    for i in range(len(data)):
        price.append(float(data[i]['Value']))

    return min(price)


wb = load_workbook(filename='FidelityTickers.xlsx',data_only=True)
ws = wb['Date']

# Read the cell values into a list of lists
dates = []
for row in ws['A1':'D244']:
    datestr=str(row[0].value)
    year=int(datestr[0:4])
    month=int(datestr[4:6])
    day=int(datestr[6:8])
    d=date(year,month,day)
    offset = BMonthEnd()
    lastday=offset.rollforward(d)
    if len(str(lastday.month))==1:
        newmonth="0"+str(lastday.month)
    else:
        newmonth=str(lastday.month)
    if len(str(lastday.day))==1:
        newday="0"+str(lastday.day)
    else:
        newday=str(lastday.day)
    newenddate=str(lastday.year)+"-"+newmonth+"-"+newday
    startdate=str(lastday.year)+"-"+newmonth+"-01"
    URL_Price = "https://lt.morningstar.com/api/rest.svc/timeseries_price/9vehuxllxs?currencyId=GBP&endDate="+newenddate+"&forwardFill=true&frequency=daily&id="+str(row[2].value)+"&idType=Morningstar&outputType=json&startDate="+startdate
    row[3].value=FindSmallestPrice(GetData(URL_Price))

wb.save('FidelityTickers.xlsx')
print("finished")