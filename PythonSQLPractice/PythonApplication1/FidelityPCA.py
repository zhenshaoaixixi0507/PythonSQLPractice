from datetime import date
from openpyxl import load_workbook
from pandas.tseries.offsets import BMonthEnd
import time
from bs4 import BeautifulSoup
import requests
import json
import os
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
import statsmodels.api as sm
import statistics

def GetTempDataArray(rng):
    final_list=[]
    for row in rng:
        mylist=[row[i].value for i in range(len(row))]
        final_list.append(mylist)
    return final_list

def GetPCAPortfolioReturn(weights,x):
    returnList={}
    for i in range(len(weights)):
        tempRet=[]
        w=weights[i]
        w=abs(w)/sum(abs(w))
        for j in range(len(x)):
            temp=0.0
            for k in range(len(x[j])):
               temp=temp+x[j][k]*w[k]
            tempRet.append(temp)

        returnList[str(i)]=tempRet
    return returnList

def SetRangeValues(rng,x):
    n=0
    for row in rng:
        row[0].value=x[n]
        n=n+1

def CalculateBetas(PortRet,x):
    factors=[[0 for i in range(len(PortRet)+1)] for j in range(len(PortRet['0']))]
    for k in range(len(PortRet)+1):
        for kk in range(len(PortRet['0'])):
            if k==0:
                factors[kk][k]=1
            else:
                factors[kk][k]=PortRet[str(k-1)][kk]

    columns = len(x[0])
    expRet=[]
    for i in range(columns):
        tempRet=[x[j][i] for j in range(len(x))]
        model = sm.OLS(tempRet, factors)
        results = model.fit()
        predictRet=results.predict()
        #expRet.append(statistics.mean(predictRet))
        expRet.append(predictRet[-1])
    return expRet

def WriteExpReturn(ws3,expRet,index1,index2):
    rng=ws3['B'+str(index1):'TG'+str(index2)]
    for i in range(len(expRet)):
        rng[0][i].value=expRet[i]

wb = load_workbook(filename='FidelityPCAAnalysis.xlsx',data_only=True)
ws = wb['Return']
ws2=wb['PCA']
ws3=wb['ExpectedReturn']
pca = PCA(n_components=10)

for i in range(212):
    index_start=2+i
    index_end=37+i
    rng_index1='B'+str(index_start)
    rnd_index2='TG'+str(index_end)
    rng=ws[rng_index1:rnd_index2]
    x=GetTempDataArray(rng)
    pca.fit_transform(x)
    weights=pca.components_
    PortRet=GetPCAPortfolioReturn(weights,x)
    expRet=CalculateBetas(PortRet,x)
    WriteExpReturn(ws3,expRet,index_start,index_start)
    SetRangeValues(ws2['B'+str(index_start):'B'+str(index_end)], PortRet['0'])
    SetRangeValues(ws2['C'+str(index_start):'C'+str(index_end)], PortRet['1'])
    SetRangeValues(ws2['D'+str(index_start):'D'+str(index_end)], PortRet['2'])
    SetRangeValues(ws2['E'+str(index_start):'E'+str(index_end)], PortRet['3'])
    SetRangeValues(ws2['F'+str(index_start):'F'+str(index_end)], PortRet['4'])
    SetRangeValues(ws2['G'+str(index_start):'G'+str(index_end)], PortRet['5'])
    SetRangeValues(ws2['H'+str(index_start):'H'+str(index_end)], PortRet['6'])
    SetRangeValues(ws2['I'+str(index_start):'I'+str(index_end)], PortRet['7'])
    SetRangeValues(ws2['J'+str(index_start):'J'+str(index_end)], PortRet['8'])
    SetRangeValues(ws2['K'+str(index_start):'K'+str(index_end)], PortRet['9'])

wb.save('FidelityPCAAnalysis.xlsx')
print('finished')

